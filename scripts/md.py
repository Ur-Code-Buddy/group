import openpyxl
import random
import copy
import math
import itertools
import os
import time
import sys



# this class takes in a text file and extracts the swimmer information from it
class InputProcessor:
    def __init__(self, file_path):
        self.file = openpyxl.load_workbook(file_path, data_only=False)["Sheet1"]
        self.timings = self.get_timings()
        self.num_teams = self.get_num_teams()

    def get_timings(self):
        dummy_timings = []

        for row in self.file.iter_rows(2, self.file.max_row):
            runner_details = [row[0].value]
            for i in range(1, 5):
                runner_details.append(self.get_seconds(row[i].value))
            dummy_timings.append(runner_details)
        return dummy_timings

    def get_num_teams(self):
        division_size = []
        index = 2
        while self.file.cell(row=index, column=8).value is not None:
            division_size.append(self.file.cell(row=index, column=8).value)
            index += 1
        return division_size

    def get_seconds(self, ms):
        if ms == "inf":
            return float("inf")
        if type(ms) is not float and type(ms) is not int:
            time = str(ms)
            if ':' in time:
                split = time.split(":")
                return int(split[len(split) - 2]) * 60 + float(split[len(split) - 1])
                #return int(split[0]) * 60 + float(split[1])
        return ms


def generate_initial_division(race_information, division_no):
    team_index = [0 for i in range(race_information.num_teams[division_no])]
    team_timings = [0 for i in range(race_information.num_teams[division_no])]
    division = [-1 for i in range(race_information.num_teams[division_no] * 4)]

    beginning_index = sum([x for x in race_information.num_teams[0: division_no]]) * 4
    swimmers = copy.deepcopy(race_information.timings[beginning_index : beginning_index + race_information.num_teams[division_no] * 4])

    race_order = [3, 4, 1, 2]
    for race_index in race_order:
        chosen_swimmers = []
        for i in range(race_information.num_teams[division_no]):
            chosen_race_timings = [swimmer[race_index] for swimmer in swimmers]
            chosen_swimmer = swimmers[chosen_race_timings.index(min(chosen_race_timings))]
            swimmers.remove(chosen_swimmer)
            chosen_swimmers.append(chosen_swimmer)

        while len(chosen_swimmers) > 0:
            teams_with_less_swimmers = []
            min_num_swimmers = min(team_index)

            for j in range(len(team_index)):
                if team_index[j] == min_num_swimmers:
                    teams_with_less_swimmers.append(j)

            max_timing = max([team_timings[x] for x in teams_with_less_swimmers])
            slow_teams = []

            for x in teams_with_less_swimmers:
                if team_timings[x] == max_timing:
                    slow_teams.append(x)

            chosen_team = random.choice(slow_teams)

            race_timings = [swimmer[race_index] for swimmer in chosen_swimmers]
            fastest_swimmer = chosen_swimmers[race_timings.index(min(race_timings))]
            chosen_swimmers.remove(fastest_swimmer)

            division[chosen_team * 4 + race_index - 1] = fastest_swimmer
            team_timings[chosen_team] += fastest_swimmer[race_index]
            team_index[chosen_team] += 1

    return division

def get_timings(division, race_information, division_no):
    team_timings = [sum([division[i * 4 + j][j + 1] for j in range(4)]) for i in range(race_information.num_teams[division_no])]
    return team_timings

def get_team_timing(team):
    return sum([team[i][i+1] for i in range(4)])

def get_optimum_timings(division, race_information, division_no):
    team_timings = []
    for i in range(race_information.num_teams[division_no]):
        team = division[i * 4: i * 4 + 4]

        best_timing = float('inf')
        for perm in (itertools.permutations(team)):
            if get_team_timing(perm) < best_timing:
                best_timing = get_team_timing(perm)

        team_timings.append(best_timing)

    return team_timings

def get_optimum_division(division, race_information, division_no):
    best_division = []
    for i in range(race_information.num_teams[division_no]):
        team = division[i * 4: i * 4 + 4]

        best_timing = float('inf')
        for perm in (itertools.permutations(team)):
            if get_team_timing(perm) < best_timing:
                best_team = perm
                best_timing = get_team_timing(perm)

        best_division += best_team

    return best_division

def get_range(division, race_information, division_no):
    team_timing = get_optimum_timings(division, race_information, division_no)
    return max(team_timing) - min(team_timing)

def fitness(division, race_information, division_no):
    range = get_range(division, race_information, division_no)
    if range == 0:
        return float('inf')
    return 1 / range

def create_neighbours(division):
    neighbours = []
    for i in range(len(division)):
        for j in range(i + 1, len(division)):
            new_neighbour = copy.deepcopy(division)
            new_neighbour[i], new_neighbour[j] = new_neighbour[j], new_neighbour[i]
            neighbours.append(new_neighbour)
    return neighbours

def main():
    # Get input and output file paths from command line arguments
    path = sys.argv[1]
    output_path = sys.argv[2]
    information = InputProcessor(str(path))

    num_competitive_runners = sum([no_of_teams for no_of_teams in information.num_teams]) * 4

    final_divisions = []
    print("starting optimization")

    for j in range(len(information.num_teams)):
        initial_division = current_division = generate_initial_division(information, j)
        num_iteration = 1000
        initial_temperature = current_temperature = 50
        cooling_rate = 0.99
        start_time = time.time()
        for i in range(num_iteration):
            current_temperature *= cooling_rate

            neighbours = create_neighbours(current_division)
            fitness_scores = [fitness(neighbour, information, j) for neighbour in neighbours]
            chosen_division = neighbours[fitness_scores.index(max(fitness_scores))]
            chosen_fitness = fitness(chosen_division, information, j)
            current_fitness = fitness(current_division, information, j)

            if chosen_fitness >= current_fitness:
                current_division = chosen_division
            else:
                probability = 1 / (1 - math.exp((chosen_fitness - current_fitness) / current_temperature))
                if random.random() < probability:
                    current_division = chosen_division

        current_division = get_optimum_division(current_division, information, j)
        final_divisions.append(current_division)

    wb = openpyxl.Workbook()
    sheet = wb.active

    races = ["Back", "Breast", "B-fly", "Crawl"]
    for index, division in enumerate(final_divisions):
        sheet.cell(row=index * 6 + 1, column=1).value = "Division " + str(index + 1)
        for k in range(5):
            if k < 4:
                sheet.cell(row=index * 6 + k + 2, column=1).value = races[k]
            for l in range(information.num_teams[index]):
                if k < 4:
                    sheet.cell(row=index * 6 + k + 2, column=l * 3 + 3).value = division[l * 4 + k][0]
                    sheet.cell(row=index * 6 + k + 2, column=l * 3 + 4).value = division[l * 4 + k][k + 1]
                else:
                    sheet.cell(row=index * 6 + k + 2, column=l * 3 + 4).value = get_timings(division, information, index)[l]

    wb.save(output_path)
    print("File generated")



main()
